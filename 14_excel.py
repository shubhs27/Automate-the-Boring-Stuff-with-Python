import openpyxl

workbook= openpyxl.load_workbook('example.xlsx')
print(type(workbook))

print(workbook.sheetnames)
sheet= workbook['Sheet1']
print(type(sheet))
print()


cell= sheet['A1']
print(cell.value)
print(sheet['B1'].value)

print(sheet.cell(row=1, column=2))
print()

for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)
print()

##################################################

wb= openpyxl.Workbook()
print(wb)

print(wb.sheetnames)
sheet= wb['Sheet']
print(sheet)
print()

print(sheet['A7'].value)
sheet['A1']= 27
sheet['A2']= 'Hello'

sheet2= wb.create_sheet()
print(wb.sheetnames)
print(sheet2.title)
sheet2.title= 'New Sheet'
print(wb.sheetnames)
print()

sheet3= wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)


wb.save('example1.xlsx')